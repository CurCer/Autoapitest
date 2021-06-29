# coding=utf-8
import openpyxl
from openpyxl import load_workbook


class Excldata:
    def __init__(self, filename):
        self.filename = filename
        self.lw = openpyxl.load_workbook(self.filename)
        self.tab = self.lw.get_sheet_by_name('Sheet1')

    # 获取行数
    def getrow(self):
        rows = self.tab.max_row
        return rows

    # 获取某个单元格的内容
    def getcellvalue(self, row, column):
        return self.tab.cell(row=row, column=column).value

    # 写入某个单元格内容
    def writecellvalue(self, row, column, cellvalue):
        try:
            self.tab.cell(row=row, column=column, value=cellvalue)
            self.lw.save(self.filename)
        except:
            self.tab.cell(row=row, column=column).value = "writefail"
            self.lw.save(self.filename)
